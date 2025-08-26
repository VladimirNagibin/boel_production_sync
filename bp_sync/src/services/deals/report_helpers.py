from datetime import datetime
from enum import Enum
from tempfile import NamedTemporaryFile
from typing import Any  # , cast

import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from pytz import UTC  # type: ignore[import-untyped]

from core.logger import logger
from models.deal_models import Deal as DealDB
from schemas.deal_schemas import DealCreate

from .enums import CreationSourceEnum, DealSourceEnum, DealTypeEnum
from .handling_helpers import identify_source


async def create_dataframe(data: list[dict[str, Any]]) -> pd.DataFrame:
    """
    Создает и сортирует DataFrame из переданных данных.

    Args:
        data: Список словарей с данными для DataFrame

    Returns:
        Отсортированный DataFrame
    """
    try:
        df = pd.DataFrame(data)
        logger.info(f"Created DataFrame with {len(df)} rows")

        df = df.sort_values(
            by=["Дата сделки", "Ответственный по сделке", "Сумма сделки"],
            ascending=[
                True,  # deal_date_create: по возрастанию (старые -> новые)
                True,  # deal_assigned_by_id: A->Я
                False,  # deal_opportunity: по убыванию (крупные сначала)
            ],
            na_position="last",  # поместить пустые значения в конец
            # inplace=True,
        )
        return df
    except Exception as e:
        logger.error(f"Error creating DataFrame: {e}")
        raise


def _get_comment(comm: str) -> Comment:
    try:
        comment = Comment(comm, "")
        comment.width = 250  # ширина в пикселях
        comment.height = 50  # высота в пикселях
        return comment
    except Exception as e:
        logger.error(f"Error creating comment: {e}")
        raise


def _setup_cell_formats(worksheet: Worksheet, df_len: int) -> None:
    """Настраивает форматы ячеек"""
    # Формулы для первой строки
    try:
        formula_row = [
            "",  # A
            '=IF(H1=0, "-",AM1/H1)',  # B
            '=IF(AM1=0, "-",AR1/AM1)',  # C
            '=IF(AR1=0, "-",AO1/AR1)',  # D
            "",  # E
            "",  # F
            "",  # G
            "=SUBTOTAL(9,H3:H" + str(df_len + 2) + ")",  # H
            *([""] * 30),  # до AM
            "=SUBTOTAL(9,AM3:AM" + str(df_len + 2) + ")",  # AM
            "",  # AN
            "=SUBTOTAL(9,AO3:AO" + str(df_len + 2) + ")",  # AO
            "",  # AP
            "",  # AQ
            "=SUBTOTAL(9,AR3:AR" + str(df_len + 2) + ")",  # AR
            *([""] * 5),  # до AW
        ]

        for col_idx, value in enumerate(formula_row, 1):
            if value:
                worksheet.cell(row=1, column=col_idx, value=value)

        # Форматирование формул
        formula_font = Font(bold=True, italic=True)
        formula_fill = PatternFill(
            start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"
        )

        for cell in worksheet[1]:
            cell.font = formula_font
            cell.fill = formula_fill

        # Процентное форматирование для формул
        for col in ["B", "C", "D"]:
            worksheet[f"{col}1"].number_format = "0.00%"

        # Форматирование сумм
        sum_format = "# ### ##0.00"
        for col in ["H", "AM", "AO", "AR"]:
            worksheet[f"{col}1"].number_format = sum_format

        worksheet["B1"].comment = _get_comment(
            "Конверсия.\nСумма сделок -> Сумма счетов"
        )
        worksheet["C1"].comment = _get_comment(
            "Конверсия.\nСумма счетов -> Сумма накладных"
        )
        worksheet["D1"].comment = _get_comment(
            "Конверсия.\nСумма накладных -> Оплаты"
        )
    except Exception as e:
        logger.error(f"Error setting up cell formats: {e}")
        raise


def _setup_column_widths(worksheet: Worksheet) -> None:
    """Настраивает ширину столбцов"""
    try:
        column_widths = {
            "A": 8,
            "B": 10,
            "C": 30,
            "D": 20,
            "E": 15,
            "F": 15,
            "G": 15,
            "H": 15,
            "I": 15,
            "J": 7,
            "K": 15,
            "L": 15,
            "M": 15,
            "N": 15,
            "O": 12,
            "P": 12,
            "Q": 12,
            "R": 10,
            "S": 6,
            "T": 10,
            "U": 12,
            "V": 8,
            "W": 10,
            "X": 30,
            "Y": 15,
            "Z": 15,
            "AA": 10,
            "AB": 10,
            "AC": 15,
            "AD": 10,
            "AE": 12,
            "AF": 8,
            "AG": 10,
            "AH": 20,
            "AI": 15,
            "AJ": 13,
            "AK": 15,
            "AL": 10,
            "AM": 15,
            "AN": 10,
            "AO": 15,
            "AP": 10,
            "AQ": 15,
            "AR": 15,
            "AS": 15,
            "AT": 15,
            "AU": 15,
            "AV": 15,
            "AW": 15,
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
    except Exception as e:
        logger.error(f"Error setting column widths: {e}")
        raise


def _format_dates(worksheet: Worksheet, df_len: int) -> None:
    """Форматирует даты в указанных колонках"""
    try:
        date_columns = ["B", "I", "W", "AG", "AT", "AW"]
        date_format = "d mmm yy"

        for col in date_columns:
            for row in range(3, df_len + 3):
                cell = worksheet[f"{col}{row}"]
                try:
                    if cell.value and isinstance(
                        cell.value, (datetime, pd.Timestamp)
                    ):
                        cell.number_format = date_format
                    elif cell.value and isinstance(cell.value, str):
                        date_value = pd.to_datetime(cell.value)
                        cell.value = date_value
                        cell.number_format = date_format
                except (ValueError, TypeError):
                    pass
    except Exception as e:
        logger.error(f"Error formatting dates: {e}")
        raise


def _format_sums(worksheet: Worksheet, df_len: int) -> None:
    """Форматирует суммы в указанных колонках"""
    try:
        sum_format = "# ### ##0.00"
        sum_columns = ["H", "K", "AM", "AO", "AR"]

        for col in sum_columns:
            for row in range(3, df_len + 3):
                cell = worksheet[f"{col}{row}"]
                try:
                    cell.number_format = sum_format
                except (ValueError, TypeError):
                    pass
    except Exception as e:
        logger.error(f"Error formatting sums: {e}")
        raise


def _setup_header_styles(worksheet: Worksheet, headers: list[str]) -> None:
    """Настраивает стили заголовков с контрастными цветами"""
    try:
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # Определяем цветовые группы для заголовков
        color_groups: dict[str, Any] = {
            "deal_info": {
                "color": "004A40",
                "columns": list(range(0, 6)),
            },  # A-F
            "company_deal_info": {
                "color": "127D6F",
                "columns": list(range(6, 11)),
            },  # H-K
            "deal_add_info": {
                "color": "004A40",
                "columns": list(range(11, 21)),
            },  # L-U
            "lead_info": {
                "color": "2C3332",
                "columns": list(range(21, 31)),
            },  # V-AE
            "invoice_info": {
                "color": "227469",
                "columns": list(range(31, 42)),
            },  # AF-AP
            "delivery_note_info": {
                "color": "243331",
                "columns": list(range(42, 46)),
            },  # AQ-AT
            "comment_info": {
                "color": "2A4A45",
                "columns": list(range(46, len(headers))),
            },  # AU-AW
        }

        # Применяем цвета к заголовкам
        for _, group in color_groups.items():
            fill = PatternFill(
                start_color=group["color"],
                end_color=group["color"],
                fill_type="solid",
            )

            for col_idx in group["columns"]:
                if col_idx < len(headers):
                    cell = worksheet.cell(row=2, column=col_idx + 1)
                    cell.font = header_font
                    cell.fill = fill
                    cell.alignment = header_alignment

        # Особые настройки для отдельных ячеек
        worksheet["J2"].font = Font(bold=True, color="FFFFFF", size=6)
        worksheet["S2"].font = Font(bold=True, color="FFFFFF", size=8)
        part_fill = PatternFill(
            start_color="5E6F6C",
            end_color="5E6F6C",
            fill_type="solid",
        )
        worksheet["M2"].fill = part_fill
        worksheet["O2"].fill = part_fill
        worksheet["Q2"].fill = part_fill
    except Exception as e:
        logger.error(f"Error setting header styles: {e}")
        raise


def _add_borders(
    worksheet: Worksheet, df_len: int, headers_count: int
) -> None:
    """Добавляет границы ко всем ячейкам"""
    try:
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in worksheet.iter_rows(
            min_row=1, max_row=df_len + 2, max_col=headers_count
        ):
            for cell in row:
                cell.border = thin_border
    except Exception as e:
        logger.error(f"Error adding borders: {e}")
        raise


async def create_excel_file(df: pd.DataFrame) -> str:
    """
    Создает временный Excel-файл с расширенным форматированием

    Args:
        df: DataFrame с данными для экспорта

    Returns:
        Путь к созданному файлу
    """
    try:
        if df.empty:
            df = pd.DataFrame({"Сообщение": ["Нет данных для отображения"]})

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            with pd.ExcelWriter(tmp.name, engine="openpyxl") as writer:
                # Записываем данные без заголовка и индекса
                df.to_excel(writer, index=False, startrow=2, header=False)

                # workbook = writer.book
                worksheet = writer.sheets["Sheet1"]
                headers = list(df.columns)
                df_len = len(df)

                # Добавляем заголовки
                headers = list(df.columns)
                for col_idx, header in enumerate(headers, 1):
                    worksheet.cell(row=2, column=col_idx, value=header)

                # Настройка различных аспектов таблицы
                _setup_cell_formats(worksheet, df_len)
                _setup_column_widths(worksheet)
                _setup_header_styles(worksheet, headers)
                _format_dates(worksheet, df_len)
                _format_sums(worksheet, df_len)
                _add_borders(worksheet, df_len, len(headers))

                # Дополнительные настройки
                worksheet.freeze_panes = "A3"
                worksheet.auto_filter.ref = (
                    f"A2:{get_column_letter(len(headers))}{df_len+2}"
                )

                # AE to left
                # AT format date

            return tmp.name
    except Exception as e:
        logger.error(f"Error creating Excel file: {e}")
        raise


def get_name(value: Any) -> str:
    """
    Безопасное получение имени из объекта с проверкой различных атрибутов

    Args:
        value: Объект для извлечения имени

    Returns:
        Строковое представление имени или пустая строка
    """
    if value is None:
        return ""

    # Проверка различных атрибутов, которые могут содержать имя
    for attr in ["full_name", "name", "title"]:
        if hasattr(value, attr) and getattr(value, attr):
            return str(getattr(value, attr))

    # Для перечислений возвращаем имя значения
    if isinstance(value, Enum):
        return value.name

    return ""


async def process_deal_row_report(deal: DealDB) -> dict[str, Any]:
    """
    Обрабатывает одну сделку для экспорта в отчет

    Args:
        deal: Объект сделки из БД

    Returns:
        Словарь с данными для строки отчета
    """
    try:
        row = await _build_data_row(deal)
        return {
            key: _convert_to_naive_datetime(value)
            for key, value in row.items()
        }
    except Exception as e:
        logger.error(f"Error processing deal {deal.id}: {e}")
        return {}


async def _build_data_row(deal: DealDB) -> dict[str, Any]:
    """
    Строит строку данных для экспорта на основе сделки

    Args:
        deal: Объект сделки из БД

    Returns:
        Словарь с данными для экспорта
    """
    try:
        deal_schema: DealCreate = deal.to_pydantic()
        lead_schema = deal.lead.to_pydantic() if deal.lead else None
        company_schema = deal.company.to_pydantic() if deal.company else None
        comments_text = ""
        if deal.timeline_comments:
            comments_ = [
                comm.comment_entity
                for comm in deal.timeline_comments
                if comm.comment_entity
            ]
            comments_text = "; ".join(comments_)
        creation_source, deal_type, deal_source = await identify_source(
            deal_b24=deal_schema,
            lead=lead_schema,
            company=company_schema,
            comments_=comments_text,
        )

        row: dict[str, Any] = {
            "ИД сделки": deal.external_id,
            "Дата сделки": deal.date_create,
            "Название сделки": deal.title,
            "Отдел": get_name(deal.assigned_user.department),
            "Ответственный по сделке": get_name(deal.assigned_user),
            "Создатель сделки": get_name(deal.created_user),
            "Клиент": get_name(deal.company),
            "Сумма сделки": deal.opportunity,
            "Дата создания клиента": (
                deal.company.date_create if deal.company else ""
            ),
            "Разница дат создания клиента и сделки": (
                (deal.date_create - deal.company.date_create).days
                if deal.company
                else ""
            ),
            "Сумма отгрузок за год": (
                deal.company.revenue if deal.company else ""
            ),
            "Тип созд. сделки": get_name(deal.creation_source),
            "Тип создания новый (авто/ручной)": (
                CreationSourceEnum.get_display_name(creation_source)
            ),
            "Источник сделки": get_name(deal.source),
            "Источник новый": DealSourceEnum.get_display_name(deal_source),
            "Тип сделки": get_name(deal.type),
            "Тип новый": DealTypeEnum.get_display_name(deal_type),
            "Стадия сделки": get_name(deal.stage),
            "Вн ном сделки": deal.origin_id,
            "ИД сайта Calltouch": deal.calltouch_site_id,
            # "deal_calltouch_call_id": deal.calltouch_call_id,
            # "deal_calltouch_request_id": deal.calltouch_request_id,
            "ИД клиента Яндекс": deal.yaclientid,
        }

        # Данные лида
        if deal.lead:
            row.update(
                {
                    "ИД лида": deal.lead.external_id,
                    "Дата лида": deal.lead.date_create,
                    "Название лида": deal.lead.title,
                    "Ответственный по лиду": get_name(deal.lead.assigned_user),
                    "Создатель лида": get_name(deal.lead.created_user),
                    "Источник лида": get_name(deal.lead.source),
                    "Тип лида": get_name(deal.lead.type),
                    "Стадия лида": get_name(deal.lead.status),
                    "ИД сайта Calltouch лид": deal.lead.calltouch_site_id,
                    "ИД клиента Яндекс лид": deal.lead.yaclientid,
                }
            )
        # Данные комментариев
        if deal.timeline_comments:
            comments: list[str] = []
            authors: set[str] = set()
            date_comm: datetime | None = None
            for comm in deal.timeline_comments:
                if comm.comment_entity:
                    comments.append(comm.comment_entity)
                author_name = get_name(comm.author)
                if author_name:
                    authors.add(author_name)
                if date_comm is None:
                    date_comm = comm.created
            row.update(
                {
                    "Комментарии из ленты": "; ".join(comments),
                    "Автор комментария": (
                        ", ".join(authors) if authors else ""
                    ),
                    "Дата комментария": date_comm,
                }
            )

        # Данные счетов
        if deal.invoices:
            invoice = deal.invoices[0]
            # print(invoice)
            row.update(
                {
                    "ИД счета": invoice.external_id,
                    "Дата счета": invoice.date_create,
                    "Название счета": invoice.title,
                    "Ответственный по счету": get_name(invoice.assigned_user),
                    "Номер счета": invoice.account_number,
                    "Компания": get_name(invoice.company),
                    "Выгружен в 1С": invoice.is_loaded,
                    "Сумма счета": invoice.opportunity,
                    "Стадия счета": get_name(invoice.invoice_stage),
                    "Оплачено по счету": sum(
                        b.amount for b in invoice.billings
                    ),
                    "Статус оплаты": (
                        "Оплачен"
                        if abs(
                            sum(b.amount for b in invoice.billings)
                            - invoice.opportunity
                        )
                        < 0.01
                        else "Частично" if any(invoice.billings) else "-"
                    ),
                }
            )

            # Данные накладных
            if invoice.delivery_notes:
                names: list[str] = []
                opportunity = 0.0
                assigned_users: set[str] = set()
                date_delivery_note = None
                for note in invoice.delivery_notes:
                    names.append(note.name)
                    opportunity += note.opportunity
                    user_name = get_name(note.assigned_user)
                    if user_name:
                        assigned_users.add(user_name)
                    if date_delivery_note is None:
                        date_delivery_note = note.date_delivery_note
                row.update(
                    {
                        "Накладная 1С": ", ".join(names),
                        "Сумма накладной": opportunity,
                        "Ответственный по накладной": (
                            ", ".join(assigned_users) if assigned_users else ""
                        ),
                        "Дата накладной": date_delivery_note,
                    }
                )
            # else:
            #    ...
            # yield invoice_row
        # else:
        #    ...
        return row
    except Exception as e:
        logger.error(f"Error building data row for deal {deal.id}: {e}")
        return {}


def _convert_to_naive_datetime(value: Any) -> Any:
    """
    Преобразует datetime с временной зоной в наивный datetime (без временной
    зоны)

    Args:
        value: Значение для преобразования

    Returns:
        Наивный datetime или оригинальное значение
    """
    if isinstance(value, datetime) and value.tzinfo is not None:
        # Конвертируем в UTC и удаляем информацию о временной зоне
        return value.astimezone(UTC).replace(tzinfo=None)
    return value
