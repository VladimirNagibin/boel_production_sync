from datetime import datetime
from tempfile import NamedTemporaryFile
from typing import Any  # , cast

import pandas as pd
from fastapi import HTTPException

# from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from core.logger import logger
from models.deal_models import Deal as DealDB

from ..base_services.base_service import BaseEntityClient
from .deal_bitrix_services import DealBitrixClient
from .deal_repository import DealRepository

# from openpyxl.worksheet.table import Table, TableStyleInfo
# from openpyxl.workbook import Workbook
# from pandas.io.excel._base import ExcelWriter


class DealClient(BaseEntityClient[DealDB, DealRepository, DealBitrixClient]):
    def __init__(
        self,
        deal_bitrix_client: DealBitrixClient,
        deal_repo: DealRepository,
    ):
        self._bitrix_client = deal_bitrix_client
        self._repo = deal_repo

    @property
    def entity_name(self) -> str:
        return "deal"

    @property
    def bitrix_client(self) -> DealBitrixClient:
        return self._bitrix_client

    @property
    def repo(self) -> DealRepository:
        return self._repo

    async def _prepare_data(
        self, start_date: datetime, end_date: datetime
    ) -> list[dict[str, Any]]:
        """Подготавливает данные для экспорта"""
        deals = await self.repo.fetch_deals(start_date, end_date)
        if not deals:
            logger.warning(f"No deals found for {start_date}-{end_date}")
            raise HTTPException(
                status_code=404, detail="No deals found in specified period"
            )

        return [await self._process_deal_row(deal) for deal in deals]

    async def _process_deal_row(self, deal: DealDB) -> dict[str, Any]:
        """Обрабатывает одну сделку для экспорта"""
        row = self.repo.build_data_row(deal)
        return {
            key: self._repo.convert_to_naive_datetime(value)
            for key, value in row.items()
        }

    async def _create_dataframe(
        self, data: list[dict[str, Any]]
    ) -> pd.DataFrame:
        df = pd.DataFrame(data)
        logger.info(f"Created DataFrame with {len(df)} rows")

        df = df.sort_values(
            by=["Дата сделки", "Ответственный по сделке", "Сумма сделки"],
            ascending=[
                True,  # deal_date_create: по возрастанию (старые -> новые)
                True,  # deal_assigned_by_id: A->Я
                False,  # deal_opportunity: по убыванию (крупные сначала)
            ],
            na_position="last",  # поместить пустые значения в конец
            # inplace=True,
        )
        return df

    def _get_comment(self, comm: str) -> Comment:
        comment = Comment(comm, "")
        comment.width = 250  # ширина в пикселях
        comment.height = 50  # высота в пикселях
        return comment

    def _setup_cell_formats(self, worksheet: Worksheet, df_len: int) -> None:
        """Настраивает форматы ячеек"""
        # Формулы для первой строки
        formula_row = [
            "",  # A
            '=IF(H1=0, "-",AM1/H1)',  # B
            '=IF(AM1=0, "-",AR1/AM1)',  # C
            '=IF(AR1=0, "-",AO1/AR1)',  # D
            "",  # E
            "",  # F
            "",  # G
            "=SUBTOTAL(9,H3:H" + str(df_len + 2) + ")",  # H
            *([""] * 30),  # до AM
            "=SUBTOTAL(9,AM3:AM" + str(df_len + 2) + ")",  # AM
            "",  # AN
            "=SUBTOTAL(9,AO3:AO" + str(df_len + 2) + ")",  # AO
            "",  # AP
            "",  # AQ
            "=SUBTOTAL(9,AR3:AR" + str(df_len + 2) + ")",  # AR
            *([""] * 5),  # до AW
        ]

        for col_idx, value in enumerate(formula_row, 1):
            if value:
                worksheet.cell(row=1, column=col_idx, value=value)

        # Форматирование формул
        formula_font = Font(bold=True, italic=True)
        formula_fill = PatternFill(
            start_color="EEEEEE", end_color="EEEEEE", fill_type="solid"
        )

        for cell in worksheet[1]:
            cell.font = formula_font
            cell.fill = formula_fill

        # Процентное форматирование для формул
        for col in ["B", "C", "D"]:
            worksheet[f"{col}1"].number_format = "0.00%"

        # Форматирование сумм
        sum_format = "# ### ##0.00"
        for col in ["H", "AM", "AO", "AR"]:
            worksheet[f"{col}1"].number_format = sum_format

        worksheet["B1"].comment = self._get_comment(
            "Конверсия.\nСумма сделок -> Сумма счетов"
        )
        worksheet["C1"].comment = self._get_comment(
            "Конверсия.\nСумма счетов -> Сумма накладных"
        )
        worksheet["D1"].comment = self._get_comment(
            "Конверсия.\nСумма накладных -> Оплаты"
        )

    def _setup_column_widths(self, worksheet: Worksheet) -> None:
        """Настраивает ширину столбцов"""
        column_widths = {
            "A": 8,
            "B": 10,
            "C": 30,
            "D": 20,
            "E": 15,
            "F": 15,
            "G": 15,
            "H": 15,
            "I": 15,
            "J": 7,
            "K": 15,
            "L": 15,
            "M": 15,
            "N": 15,
            "O": 12,
            "P": 12,
            "Q": 12,
            "R": 10,
            "S": 6,
            "T": 10,
            "U": 12,
            "V": 8,
            "W": 10,
            "X": 30,
            "Y": 15,
            "Z": 15,
            "AA": 10,
            "AB": 10,
            "AC": 15,
            "AD": 10,
            "AE": 12,
            "AF": 8,
            "AG": 10,
            "AH": 20,
            "AI": 15,
            "AJ": 13,
            "AK": 15,
            "AL": 10,
            "AM": 15,
            "AN": 10,
            "AO": 15,
            "AP": 10,
            "AQ": 15,
            "AR": 15,
            "AS": 15,
            "AT": 15,
            "AU": 15,
            "AV": 15,
            "AW": 15,
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

    def _format_dates(self, worksheet: Worksheet, df_len: int) -> None:
        """Форматирует даты в указанных колонках"""
        date_columns = ["B", "I", "W", "AG", "AT", "AW"]
        date_format = "d mmm yy"

        for col in date_columns:
            for row in range(3, df_len + 3):
                cell = worksheet[f"{col}{row}"]
                try:
                    if cell.value and isinstance(
                        cell.value, (datetime, pd.Timestamp)
                    ):
                        cell.number_format = date_format
                    elif cell.value and isinstance(cell.value, str):
                        date_value = pd.to_datetime(cell.value)
                        cell.value = date_value
                        cell.number_format = date_format
                except (ValueError, TypeError):
                    pass

    def _format_sums(self, worksheet: Worksheet, df_len: int) -> None:
        """Форматирует суммы в указанных колонках"""
        sum_format = "# ### ##0.00"
        sum_columns = ["H", "K", "AM", "AO", "AR"]

        for col in sum_columns:
            for row in range(3, df_len + 3):
                cell = worksheet[f"{col}{row}"]
                try:
                    cell.number_format = sum_format
                except (ValueError, TypeError):
                    pass

    def _setup_header_styles(
        self, worksheet: Worksheet, headers: list[str]
    ) -> None:
        """Настраивает стили заголовков с контрастными цветами"""
        header_font = Font(bold=True, color="FFFFFF", size=10)
        header_alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # Определяем цветовые группы для заголовков
        color_groups: dict[str, Any] = {
            "deal_info": {
                "color": "004A40",
                "columns": list(range(0, 6)),
            },  # A-F
            "company_deal_info": {
                "color": "127D6F",
                "columns": list(range(6, 11)),
            },  # H-K
            "deal_add_info": {
                "color": "004A40",
                "columns": list(range(11, 21)),
            },  # L-U
            "lead_info": {
                "color": "2C3332",
                "columns": list(range(21, 31)),
            },  # V-AE
            "invoice_info": {
                "color": "227469",
                "columns": list(range(31, 42)),
            },  # AF-AP
            "delivery_note_info": {
                "color": "243331",
                "columns": list(range(42, 46)),
            },  # AQ-AT
            "comment_info": {
                "color": "2A4A45",
                "columns": list(range(46, len(headers))),
            },  # AU-AW
        }

        # Применяем цвета к заголовкам
        for _, group in color_groups.items():
            fill = PatternFill(
                start_color=group["color"],
                end_color=group["color"],
                fill_type="solid",
            )

            for col_idx in group["columns"]:
                if col_idx < len(headers):
                    cell = worksheet.cell(row=2, column=col_idx + 1)
                    cell.font = header_font
                    cell.fill = fill
                    cell.alignment = header_alignment

        # Особые настройки для отдельных ячеек
        worksheet["J2"].font = Font(bold=True, color="FFFFFF", size=6)
        worksheet["S2"].font = Font(bold=True, color="FFFFFF", size=8)
        part_fill = PatternFill(
            start_color="5E6F6C",
            end_color="5E6F6C",
            fill_type="solid",
        )
        worksheet["M2"].fill = part_fill
        worksheet["O2"].fill = part_fill
        worksheet["Q2"].fill = part_fill

    def _add_borders(
        self, worksheet: Worksheet, df_len: int, headers_count: int
    ) -> None:
        """Добавляет границы ко всем ячейкам"""
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for row in worksheet.iter_rows(
            min_row=1, max_row=df_len + 2, max_col=headers_count
        ):
            for cell in row:
                cell.border = thin_border

    async def _create_excel_file(self, df: pd.DataFrame) -> str:
        """Создает временный Excel-файл с расширенным форматированием"""
        if df.empty:
            df = pd.DataFrame({"Сообщение": ["Нет данных для отображения"]})

        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
            with pd.ExcelWriter(tmp.name, engine="openpyxl") as writer:
                # Записываем данные без заголовка и индекса
                df.to_excel(writer, index=False, startrow=2, header=False)

                # workbook = writer.book
                worksheet = writer.sheets["Sheet1"]
                headers = list(df.columns)
                df_len = len(df)

                # Добавляем заголовки
                headers = list(df.columns)
                for col_idx, header in enumerate(headers, 1):
                    worksheet.cell(row=2, column=col_idx, value=header)

                # Настройка различных аспектов таблицы
                self._setup_cell_formats(worksheet, df_len)
                self._setup_column_widths(worksheet)
                self._setup_header_styles(worksheet, headers)
                self._format_dates(worksheet, df_len)
                self._format_sums(worksheet, df_len)
                self._add_borders(worksheet, df_len, len(headers))

                # Дополнительные настройки
                worksheet.freeze_panes = "A3"
                worksheet.auto_filter.ref = (
                    f"A2:{get_column_letter(len(headers))}{df_len+2}"
                )

                # AE to left
                # AT format date

            return tmp.name

    async def export_deals_to_excel(
        self,
        start_date: datetime,
        end_date: datetime,
    ) -> str:
        """Основной метод экспорта сделок в Excel"""
        try:
            logger.info(
                f"Exporting deals to excel from {start_date} to {end_date}"
            )
            data = await self._prepare_data(start_date, end_date)
            df = await self._create_dataframe(data)
            return await self._create_excel_file(df)
        except Exception as e:
            logger.error(f"Export failed: {str(e)}", exc_info=True)
            raise HTTPException(
                status_code=500, detail=f"Export error: {str(e)}"
            ) from e
